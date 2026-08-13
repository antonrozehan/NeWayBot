import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from config import EXCEL_PATH

class ExcelExporter:
    def __init__(self):
        self.file_path = EXCEL_PATH

    def is_free_day(self, time_text):
        """Проверяет, является ли день свободным"""
        free_keywords = ['wolne', 'nie mogę', 'off', 'free', 'nie moge', 'wolny']
        time_lower = time_text.lower()
        return any(keyword in time_lower for keyword in free_keywords)

    def export_schedules_to_excel(self, schedules, week_start, confirmed_shifts=None):
        """Экспорт графиков в Excel - группировка по дням (только рабочие дни)"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grafik"
        
        # Стили
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        day_font = Font(bold=True, size=11)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # ЗАГОЛОВОК
        ws.merge_cells('A1:D1')
        ws['A1'] = "📋 GRAFIK PRACOWNIKÓW"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Tydzień: {week_start}"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:D3')
        ws['A3'] = f"Data: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].alignment = center_alignment
        ws['A3'].font = Font(size=10)
        
        # ЗАГОЛОВКИ СТОЛБЦОВ
        headers = ['Dzień', 'Pracownik', 'Godziny', 'Username']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # СОБИРАЕМ ДАННЫЕ ПО ДНЯМ (только рабочие дни)
        days_order = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
        days_data = {day: [] for day in days_order}
        
        if schedules:
            for s in schedules:
                full_name = s.get('full_name', 'Неизвестно')
                username = s.get('username', '')
                schedule_text = s.get('schedule_text', '')
                
                # Парсим график сотрудника
                lines = schedule_text.strip().split('\n')
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Определяем день и время
                    for day in days_order:
                        if day.lower() in line.lower():
                            # Извлекаем время
                            time_part = ""
                            if ':' in line:
                                parts = line.split(':', 1)
                                if len(parts) > 1:
                                    time_part = parts[1].strip()
                                else:
                                    for d in days_order:
                                        if d.lower() in line.lower():
                                            time_part = line.replace(d, '').strip()
                                            break
                            else:
                                for d in days_order:
                                    if d.lower() in line.lower():
                                        time_part = line.replace(d, '').strip()
                                        break
                            
                            if not time_part:
                                time_part = "cały dzień"
                            
                            # ПРОВЕРКА: если это свободный день - пропускаем (не добавляем в таблицу)
                            if self.is_free_day(time_part):
                                break
                            
                            # Добавляем только рабочие дни
                            if day in days_data:
                                days_data[day].append({
                                    'name': full_name,
                                    'time': time_part,
                                    'username': f"@{username}" if username else 'brak username'
                                })
                            break
        
        # ЗАПОЛНЯЕМ ТАБЛИЦУ (только дни, где есть сотрудники)
        row = 6
        
        for day in days_order:
            if day not in days_data or not days_data[day]:
                continue
            
            employees = days_data[day]
            
            # Записываем день
            start_row = row
            for idx, emp in enumerate(employees):
                if idx == 0:
                    # Первая строка - пишем день
                    cell = ws.cell(row=row, column=1, value=day)
                    cell.font = day_font
                    cell.fill = day_fill
                    cell.border = border
                    cell.alignment = center_alignment
                else:
                    ws.cell(row=row, column=1, value="").border = border
                
                # Сотрудник
                ws.cell(row=row, column=2, value=emp['name']).border = border
                ws.cell(row=row, column=2).alignment = left_alignment
                
                # Время
                ws.cell(row=row, column=3, value=emp['time']).border = border
                ws.cell(row=row, column=3).alignment = center_alignment
                
                # Username
                ws.cell(row=row, column=4, value=emp['username']).border = border
                ws.cell(row=row, column=4).alignment = center_alignment
                
                row += 1
            
            # Объединяем ячейки для дня (если больше одного сотрудника)
            if len(employees) > 1:
                ws.merge_cells(f'A{start_row}:A{row-1}')
        
        # АВТОМАТИЧЕСКАЯ ШИРИНА
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        wb.save(self.file_path)
        return self.file_path

    def export_free_employees_to_excel(self, free_employees_data, day, week_start):
        """Экспорт списка свободных сотрудников"""
        
        free_file_path = self.file_path.replace('schedules.xlsx', f'free_employees_{day}_{week_start}.xlsx')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Wolni {day}"
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f"📋 WOLNI PRACOWNICY - {day}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Tydzień: {week_start}"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:D3')
        ws['A3'] = f"Data: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].alignment = center_alignment
        ws['A3'].font = Font(size=10)
        
        headers = ['№', 'Imię i Nazwisko', 'Username', 'Godziny']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        row = 6
        for idx, emp in enumerate(free_employees_data, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center_alignment
            
            ws.cell(row=row, column=2, value=emp.get('full_name', 'Неизвестно')).border = border
            ws.cell(row=row, column=2).alignment = left_alignment
            
            username = emp.get('username', 'brak username')
            ws.cell(row=row, column=3, value=f"@{username}").border = border
            ws.cell(row=row, column=3).alignment = center_alignment
            
            ws.cell(row=row, column=4, value=emp.get('time', 'Wolny')).border = border
            ws.cell(row=row, column=4).alignment = center_alignment
            
            row += 1
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        
        wb.save(free_file_path)
        return free_file_path

    def update_excel_with_new_schedule(self, schedule_data, week_start):
        """Обновить Excel при новом графике"""
        
        from database import Database
        db = Database()
        schedules = []
        
        try:
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            schedules = loop.run_until_complete(db.get_all_schedules_with_users())
        except Exception as e:
            print(f"Ошибка получения данных: {e}")
            schedules = []
        
        if not schedules:
            schedules = [{
                'user_id': schedule_data['user_id'],
                'full_name': schedule_data['full_name'],
                'username': schedule_data.get('username', ''),
                'schedule_text': schedule_data['schedule_text'],
                'created_at': schedule_data.get('created_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            }]
        else:
            found = False
            for i, s in enumerate(schedules):
                if s['user_id'] == schedule_data['user_id']:
                    schedules[i] = {
                        'user_id': schedule_data['user_id'],
                        'full_name': schedule_data['full_name'],
                        'username': schedule_data.get('username', ''),
                        'schedule_text': schedule_data['schedule_text'],
                        'created_at': schedule_data.get('created_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    }
                    found = True
                    break
            
            if not found:
                schedules.append({
                    'user_id': schedule_data['user_id'],
                    'full_name': schedule_data['full_name'],
                    'username': schedule_data.get('username', ''),
                    'schedule_text': schedule_data['schedule_text'],
                    'created_at': schedule_data.get('created_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                })
        
        # Получаем подтвержденные смены (для цветовой индикации)
        confirmed_shifts = []
        try:
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            confirmed_shifts = loop.run_until_complete(db.get_all_confirmed_shifts())
        except:
            confirmed_shifts = []
        
        return self.export_schedules_to_excel(schedules, week_start, confirmed_shifts)
    def export_not_assigned_to_excel(self, not_assigned_employees, week_start):
        """Экспорт списка сотрудников, которые ещё не получили назначения"""
        
        not_assigned_file_path = self.file_path.replace('schedules.xlsx', f'not_assigned_{week_start}.xlsx')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Не назначенные"
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.merge_cells('A1:E1')
        ws['A1'] = f"📋 СОТРУДНИКИ БЕЗ НАЗНАЧЕНИЙ"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Тиждень: {week_start}"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].alignment = center_alignment
        ws['A3'].font = Font(size=10)
        
        # Заголовки
        headers = ['№', 'Сотрудник', 'Username', 'График', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        row = 6
        for idx, emp in enumerate(not_assigned_employees, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center_alignment
            
            ws.cell(row=row, column=2, value=emp.get('full_name', 'Неизвестно')).border = border
            ws.cell(row=row, column=2).alignment = left_alignment
            
            username = emp.get('username', 'brak username')
            ws.cell(row=row, column=3, value=f"@{username}").border = border
            ws.cell(row=row, column=3).alignment = center_alignment
            
            # График (первые 3 строки)
            schedule_text = emp.get('schedule_text', '')
            lines = schedule_text.strip().split('\n')[:3]
            schedule_display = '\n'.join(lines)
            if len(schedule_text.strip().split('\n')) > 3:
                schedule_display += '\n...'
            
            ws.cell(row=row, column=4, value=schedule_display).border = border
            ws.cell(row=row, column=4).alignment = left_alignment
            
            # Статус
            status_cell = ws.cell(row=row, column=5, value="❌ Не назначен")
            status_cell.border = border
            status_cell.alignment = center_alignment
            status_cell.fill = warning_fill
            status_cell.font = Font(bold=True)
            
            row += 1
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 18
        
        wb.save(not_assigned_file_path)
        return not_assigned_file_path