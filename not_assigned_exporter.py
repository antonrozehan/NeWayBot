import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

class NotAssignedExporter:
    def __init__(self):
        pass

    def export_not_assigned_to_excel(self, not_assigned_employees, week_start, file_path):
        """Экспорт списка сотрудников, которые не назначены на определенные дни - группировка по дням"""
        
        not_assigned_file_path = file_path.replace('schedules.xlsx', f'not_assigned_{week_start}.xlsx')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Не назначенные"
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        day_font = Font(bold=True, size=11)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # ЗАГОЛОВОК
        ws.merge_cells('A1:E1')
        ws['A1'] = f"📋 СОТРУДНИКИ БЕЗ НАЗНАЧЕНИЙ"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Тиждень: {week_start}"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].alignment = center_alignment
        ws['A3'].font = Font(size=10)
        
        # ЗАГОЛОВКИ (поменяли местами: Время в графике теперь на 3 месте, Username на 4)
        headers = ['День', 'Сотрудник', 'Время в графике', 'Username', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # СОБИРАЕМ ДАННЫЕ ПО ДНЯМ
        days_order = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
        days_data = {day: [] for day in days_order}
        
        # Заполняем данные по дням
        for emp in not_assigned_employees:
            missing_dates = emp.get('missing_dates', [])
            employee_days = emp.get('employee_days', [])
            
            # Если есть пропущенные дни
            if missing_dates:
                for m in missing_dates:
                    day = m['day']
                    if day in days_data:
                        days_data[day].append({
                            'full_name': emp.get('full_name', 'Неизвестно'),
                            'username': emp.get('username', ''),
                            'time': m['time'],
                            'date': m['date']
                        })
            elif employee_days:
                # Если сотрудник вообще не назначен - все его дни считаются пропущенными
                for emp_day in employee_days:
                    day = emp_day['day']
                    if day in days_data:
                        days_data[day].append({
                            'full_name': emp.get('full_name', 'Неизвестно'),
                            'username': emp.get('username', ''),
                            'time': emp_day['time'],
                            'date': emp_day['date']
                        })
        
        # ЗАПОЛНЯЕМ ТАБЛИЦУ - ГРУППИРОВКА ПО ДНЯМ
        row = 6
        
        for day in days_order:
            if day not in days_data or not days_data[day]:
                continue
            
            employees = days_data[day]
            
            # Записываем день (объединяем строки)
            start_row = row
            for idx, emp in enumerate(employees):
                if idx == 0:
                    # Первая строка - пишем день с датой
                    date_str = emp.get('date', '')
                    day_display = f"{day} ({date_str})" if date_str else day
                    cell = ws.cell(row=row, column=1, value=day_display)
                    cell.font = day_font
                    cell.fill = day_fill
                    cell.border = border
                    cell.alignment = center_alignment
                else:
                    ws.cell(row=row, column=1, value="").border = border
                
                # Сотрудник
                ws.cell(row=row, column=2, value=emp['full_name']).border = border
                ws.cell(row=row, column=2).alignment = left_alignment
                
                # Время в графике (с подсветкой) - теперь на 3 месте
                time_cell = ws.cell(row=row, column=3, value=emp['time'])
                time_cell.border = border
                time_cell.alignment = center_alignment
                time_cell.fill = warning_fill
                time_cell.font = Font(bold=True)
                
                # Username - теперь на 4 месте
                username = emp.get('username', 'brak username')
                ws.cell(row=row, column=4, value=f"@{username}" if username else 'brak username').border = border
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # Статус
                status_cell = ws.cell(row=row, column=5, value="❌ Не назначен")
                status_cell.border = border
                status_cell.alignment = center_alignment
                status_cell.fill = warning_fill
                status_cell.font = Font(bold=True)
                
                row += 1
            
            # Объединяем ячейки для дня (если больше одного сотрудника)
            if len(employees) > 1:
                ws.merge_cells(f'A{start_row}:A{row-1}')
        
        # АВТОМАТИЧЕСКАЯ ШИРИНА
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
        
        wb.save(not_assigned_file_path)
        return not_assigned_file_path